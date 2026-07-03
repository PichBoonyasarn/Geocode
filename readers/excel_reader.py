"""
Excel reader with two-pass coordinate extraction:
  Pass 1 — Spatial label scan: find 緯度/経度 label cells and parse the
            adjacent numeric value as compact DMS (DDMMSS.S / DDDMMSS.S),
            which is the standard format used in Japanese government inspection
            forms (e.g. 354242.4 → 35°42'42.4"N).
  Pass 2 — Full text dump so coord_parser can try its generic patterns.
"""

from openpyxl import load_workbook

# Keywords that indicate a latitude cell label
_LAT_KEYWORDS = {"緯度", "北緯", "緯     度", "緯　　　度"}
# Keywords that indicate a longitude cell label
_LON_KEYWORDS = {"経度", "東経", "経     度", "経　　　度"}

# Japan bounding box in compact DMS
_LAT_COMPACT_MIN, _LAT_COMPACT_MAX = 240000, 465959   # 24° – 46°N
_LON_COMPACT_MIN, _LON_COMPACT_MAX = 1230000, 1545959  # 123° – 154°E


def extract_text_and_images(xlsx_path: str, max_pages: int = 2) -> tuple[str, list[bytes]]:
    """
    Extract coordinates and full cell text from an Excel file.
    Returns (text, []) where text either starts with 'LAT: / LON:' lines
    (from spatial label detection) or is a flat cell dump for generic parsing.
    """
    # read_only=False needed to access cell neighbours by coordinate
    wb = load_workbook(xlsx_path, data_only=True)

    # Pass 1: spatial label scan
    coords = _find_coords_by_label(wb)

    # Pass 2: flat text dump (always included so generic patterns still run)
    texts: list[str] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for val in row:
                if val is not None:
                    s = str(val).strip()
                    if s:
                        texts.append(s)

    wb.close()

    if coords:
        lat, lon = coords
        # Prepend in a format coord_parser will definitely recognise
        header = f"緯度: {lat:.6f}\n経度: {lon:.6f}\n"
        return header + "\n".join(texts), []

    return "\n".join(texts), []


# ─────────────────────────────────────────────────────────────────────────────
# Spatial label detection helpers
# ─────────────────────────────────────────────────────────────────────────────

def _normalise_label(s: str) -> str:
    """Strip whitespace variants for fuzzy keyword matching."""
    return s.replace("　", "").replace(" ", "").replace("　", "").strip()


def _find_coords_by_label(wb) -> tuple[float, float] | None:
    """
    Scan all sheets for cells whose text matches a lat/lon keyword,
    then grab the numeric value from the nearest non-empty cell to the right.
    Returns (decimal_lat, decimal_lon) or None.
    """
    for ws in wb.worksheets:
        lat_val: float | None = None
        lon_val: float | None = None

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                label = _normalise_label(str(cell.value))

                if any(_normalise_label(k) in label for k in _LAT_KEYWORDS):
                    raw = _adjacent_number(ws, cell)
                    if raw is not None:
                        lat_val = _compact_dms_to_decimal(raw, is_lon=False)

                elif any(_normalise_label(k) in label for k in _LON_KEYWORDS):
                    raw = _adjacent_number(ws, cell)
                    if raw is not None:
                        lon_val = _compact_dms_to_decimal(raw, is_lon=True)

                if lat_val is not None and lon_val is not None:
                    return lat_val, lon_val

    return None


def _adjacent_number(ws, cell) -> float | None:
    """Return the first numeric value found in the 5 cells to the right of cell."""
    for col_offset in range(1, 6):
        try:
            adj = ws.cell(row=cell.row, column=cell.column + col_offset)
        except Exception:
            break
        if adj.value is None:
            continue
        try:
            num = float(adj.value)
            return num
        except (ValueError, TypeError):
            continue
    return None


def _compact_dms_to_decimal(value: float, is_lon: bool) -> float | None:
    """
    Convert Japanese government compact-DMS number to decimal degrees.
    Latitude format  : DDMMSS.S  (e.g. 354242.4  → 35°42'42.4"N)
    Longitude format : DDDMMSS.S (e.g. 1394307.7 → 139°43'07.7"E)
    """
    lo, hi = (_LON_COMPACT_MIN, _LON_COMPACT_MAX) if is_lon else (_LAT_COMPACT_MIN, _LAT_COMPACT_MAX)
    if not (lo <= value <= hi):
        return None

    int_part = int(value)
    frac_part = value - int_part          # fractional seconds

    if is_lon:
        ddd = int_part // 10000
        mm  = (int_part % 10000) // 100
        ss  = int_part % 100 + frac_part
    else:
        dd  = int_part // 10000
        mm  = (int_part % 10000) // 100
        ss  = int_part % 100 + frac_part
        ddd = dd  # reuse variable name for uniform return

    degrees = ddd + mm / 60.0 + ss / 3600.0
    return round(degrees, 6)
